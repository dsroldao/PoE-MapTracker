import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import time
import threading
import os
import sys
import subprocess 
import ctypes 
from datetime import datetime

# --- NEW DEPENDENCIES FOR TRAY ICON ---
try:
    import pystray
    from PIL import Image, ImageDraw
    HAS_TRAY = True
except ImportError:
    HAS_TRAY = False
    print("Tray icon libraries missing. Install: pip install pystray Pillow")

# Optional dependency check
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ==========================================
#               CONFIGURATION
# ==========================================

# --- App Info ---
APP_VERSION = "1.0.7"

# --- Logic Settings ---
SAVE_DELAY = 15  # Seconds to wait after leaving a map before saving the run
GAME_PROCESSES = ["PathOfExile.exe", "PathOfExile_x64.exe", "PathOfExileSteam.exe"]

# --- Visual Settings (Modern Theme) ---
THEME = {
    "bg": "#2f3136",
    "title_bg": "#202225",
    "text": "#ffffff",
    "accent": "#d4af37",
    "timer_idle": "#b9bbbe",
    "death": "#ed4245",
    "border": "#202225",
    "input_bg": "#40444b",
    "hover_close": "#ed4245",
    "hover_mode": "#d4af37"
}

FONTS = {
    "main": ("Segoe UI", 12),        
    "timer": ("Segoe UI", 14, "bold"), 
    "title": ("Segoe UI", 10, "bold"), 
    "badge": ("Segoe UI", 10, "bold")  
}

# --- Game Logic Settings ---
# List of Safe Zones where the timer should pause automatically
SAFE_ZONES = [
    "Lioneye's Watch", "The Forest Encampment", "The Sarn Encampment", 
    "Highgate", "Overseer's Tower", "The Bridge Encampment", 
    "Oriath", "Karui Shores", "Hideout", "Kingsmarch", 
    "The Rogue Harbour", "The Menagerie", "Mine Encampment", 
    "Aspirant's Plaza", "The Forbidden Sanctum", "Azurite Mine",
    "Monastery of the Keepers"
]

# List of Tier 17 Maps (To append T17 tag)
TIER_17_MAPS = [
    "Abomination", 
    "Citadel", 
    "Fortress", 
    "Sanctuary", 
    "Ziggurat"
]

# Mechanics Configuration
# Maps chat triggers to mechanic names, colors, and badges
MECHANICS_CONFIG = {
    # Delirium
    "The Strange Voice": {"name": "Delirium",      "color": "#A9A9A9", "badge": "D"},
    "Strange Voice":     {"name": "Delirium",      "color": "#A9A9A9", "badge": "D"},
    
    # Ultimatum
    "The Trialmaster":   {"name": "Ultimatum",     "color": "#FF4444", "badge": "U"},
    "Trialmaster":       {"name": "Ultimatum",     "color": "#FF4444", "badge": "U"},
    
    # Blight
    "Sister Cassia":     {"name": "Blight",        "color": "#FFFACD", "badge": "B"},
    
    # Incursion
    "Alva":              {"name": "Incursion",     "color": "#FFA500", "badge": "A"},
    
    # Beasts
    "Einhar":            {"name": "Beasts",        "color": "#D2691E", "badge": "Ei"},
    
    # Delve
    "Niko":              {"name": "Delve",         "color": "#4488FF", "badge": "N"},
    
    # Betrayal
    "Jun":               {"name": "Betrayal",      "color": "#90EE90", "badge": "Sy"},
    "Interrogate":       {"name": "Betrayal",      "color": "#90EE90", "badge": "Sy"},
    
    # Maven
    "The Envoy":         {"name": "Maven",         "color": "#FF00FF", "badge": "M"},
    "The Maven":         {"name": "Maven",         "color": "#FF00FF", "badge": "M"},
    
    # Expedition
    "Dannig":            {"name": "Expedition",    "color": "#00FFFF", "badge": "E"},
    "Tujen":             {"name": "Expedition",    "color": "#00FFFF", "badge": "E"},
    "Rog":               {"name": "Expedition",     "color": "#00FFFF", "badge": "E"},
    "Gwennen":           {"name": "Expedition",    "color": "#00FFFF", "badge": "E"},
    
    # Sanctum
    "Divinia":           {"name": "Sanctum",       "color": "#FFD700", "badge": "S"},
    
    # Nameless Seer
    "Nameless Seer":     {"name": "Nameless Seer", "color": "#C8A2C8", "badge": "NS"},

    # Eagon (Memory Tear mechanic)
    "Eagon Caeserius":   {"name": "Eagon",         "color": "#9370DB", "badge": "Eg"}, 
    "Eagon":             {"name": "Eagon",         "color": "#9370DB", "badge": "Eg"},

    # Harvest (Oshabi)
    "Oshabi":            {"name": "Harvest",       "color": "#7FFFD4", "badge": "H"},
    "Oshabi, Avatar of the Grove": {"name": "Harvest", "color": "#7FFFD4", "badge": "H"},

    # Zana (Tier 16.5)
    "Zana, The Originator": {"name": "Tier 16.5",   "color": "#DA70D6", "badge": "T16.5"}
}

class PoEOverlay(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # FIX: Set App ID immediately for Taskbar Icon
        try:
            # App ID for Windows Taskbar grouping
            myappid = 'dsroldao.poemaptracker.v1.0.7'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except: pass

        self._check_single_instance()
        self._check_dependencies()
        self._init_variables() 
        self._create_start_menu_shortcut() 
        self._init_window() 
        self._setup_ui()
        
        if HAS_TRAY:
            threading.Thread(target=self._setup_tray_icon, daemon=True).start()

        self.after(500, self._init_log_search)
        self.after(2000, self._monitor_game_process) 
        self.after(200, self._monitor_focus_loop) 
        
        self.after(200, self._force_taskbar_icon)

    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller/Nuitka """
        search_paths = []
        try:
            search_paths.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path))
        except: pass
        if hasattr(sys, '_MEIPASS'):
            search_paths.append(os.path.join(sys._MEIPASS, relative_path))
        try:
            search_paths.append(os.path.join(os.path.dirname(sys.executable), relative_path))
        except: pass
        search_paths.append(os.path.join(os.getcwd(), relative_path))

        for path in search_paths:
            if os.path.exists(path):
                return path
        return relative_path

    def _check_single_instance(self):
        try:
            self.mutex = ctypes.windll.kernel32.CreateMutexW(None, True, "Global\\PoE_MapTracker_Instance")
            last_error = ctypes.windll.kernel32.GetLastError()
            if last_error == 183: 
                messagebox.showwarning("Already Running", "PoE Map Tracker is already open.")
                sys.exit()
        except: pass 

    def _check_dependencies(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Dependency", "To generate Excel files, install openpyxl:\n\npip install openpyxl")
            sys.exit()

    def _init_window(self):
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("dark-blue")
        
        self.title(f"PoE Map Tracker v{APP_VERSION}")
        self.geometry("340x50+100+100")
        self.overrideredirect(True) 
        self.attributes("-topmost", True) 
        self.attributes("-alpha", 0.95)
        self.configure(fg_color=THEME["bg"])
        
        icon_path = self.resource_path("icon.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(bitmap=icon_path)
                self.iconbitmap(default=icon_path)
            except Exception as e:
                print(f"Icon load error: {e}")

        self.main_container = ctk.CTkFrame(self, fg_color=THEME["bg"], border_width=1, border_color=THEME["border"], corner_radius=0)
        self.main_container.pack(fill="both", expand=True)

    def _force_taskbar_icon(self):
        try:
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20) 
            style = style | 0x00040000 
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style)
            self.withdraw()
            self.after(10, self.deiconify)
        except: pass

    def _init_variables(self):
        self.status = "idle" 
        self.current_map = "Starting..."
        self.start_time = 0
        self.elapsed = 0
        self.cooldown_start = 0
        self.deaths = 0
        self.mechanics_found = []
        self.tier_var = tk.StringVar(value="16")
        self.running = True
        self.is_compact = False
        self.pending_runs = []
        self.game_not_found_count = 0 
        self.is_paused = False
        self.pause_timestamp = 0
        self.latest_safe_zone = "Hideout" # Stores the last visited safe zone

        if getattr(sys, 'frozen', False) or "__compiled__" in globals():
            self.app_dir = os.path.dirname(sys.executable)
        else:
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.history_dir = os.path.join(self.app_dir, "Map History")
        self._ensure_history_folder()
            
        self.config_file = os.path.join(self.app_dir, "config.txt")
        self.excel_file = os.path.join(self.history_dir, "map_history.xlsx")

    def _ensure_history_folder(self):
        if not os.path.exists(self.history_dir):
            try: os.makedirs(self.history_dir)
            except: pass

    def _create_start_menu_shortcut(self):
        if not getattr(sys, 'frozen', False): return 
        try:
            app_name = "PoE Map Tracker"
            start_menu = os.path.join(os.environ["APPDATA"], "Microsoft", "Windows", "Start Menu", "Programs")
            shortcut_path = os.path.join(start_menu, f"{app_name}.lnk")
            target_exe = sys.executable
            icon_path = self.resource_path("icon.ico")
            
            if not os.path.exists(shortcut_path):
                icon_cmd = f'$Shortcut.IconLocation = "{icon_path}"' if os.path.exists(icon_path) else ""
                ps_script = f"""
                $WshShell = New-Object -comObject WScript.Shell
                $Shortcut = $WshShell.CreateShortcut("{shortcut_path}")
                $Shortcut.TargetPath = "{target_exe}"
                $Shortcut.WorkingDirectory = "{self.app_dir}"
                $Shortcut.Description = "Automated Map Tracker for Path of Exile"
                {icon_cmd}
                $Shortcut.Save()
                """
                subprocess.run(["powershell", "-Command", ps_script], check=True, creationflags=subprocess.CREATE_NO_WINDOW)
        except: pass

    def _setup_tray_icon(self):
        if not HAS_TRAY: return
        icon_path = self.resource_path("icon.ico")
        if os.path.exists(icon_path):
            image = Image.open(icon_path)
        else:
            image = Image.new('RGB', (64, 64), color=(212, 175, 55))
            d = ImageDraw.Draw(image)
            d.rectangle([16,16,48,48], fill=(47, 49, 54))
        def on_quit(icon, item):
            icon.stop()
            self.quit()
        def on_show(icon, item):
            self.after(0, self.deiconify)
            self.after(0, lambda: self.attributes("-topmost", True))
        menu = pystray.Menu(
            pystray.MenuItem("Show Tracker", on_show, default=True),
            pystray.MenuItem("Exit", on_quit)
        )
        self.tray_icon = pystray.Icon("PoE Map Tracker", image, "PoE Map Tracker", menu)
        self.tray_icon.run()

    def _setup_ui(self):
        self.title_bar = ctk.CTkFrame(self.main_container, height=16, fg_color=THEME["title_bg"], corner_radius=0)
        self.title_bar.pack(fill="x", side="top")
        
        self.lbl_title = ctk.CTkLabel(self.title_bar, text=f" PoE Map Tracker v{APP_VERSION}", 
                                      font=FONTS["title"], text_color=THEME["timer_idle"])
        self.lbl_title.pack(side="left", padx=5)

        controls = ctk.CTkFrame(self.title_bar, fg_color="transparent")
        controls.pack(side="right", fill="y")

        self.btn_mode = ctk.CTkButton(controls, text="M", width=20, height=16, 
                                      fg_color="transparent", hover_color=THEME["hover_mode"],
                                      text_color=THEME["timer_idle"], font=FONTS["title"],
                                      command=self._toggle_compact_mode)
        self.btn_mode.pack(side="left")

        btn_close = ctk.CTkButton(controls, text="X", width=20, height=16,
                                  fg_color="transparent", hover_color=THEME["hover_close"],
                                  text_color=THEME["timer_idle"], font=FONTS["title"],
                                  command=self.quit) 
        btn_close.pack(side="left")

        self.content = ctk.CTkFrame(self.main_container, fg_color="transparent", corner_radius=0)
        self.content.pack(fill="both", expand=True, padx=5)

        self.lbl_map = ctk.CTkLabel(self.content, text=self.current_map, font=FONTS["main"], text_color=THEME["text"])
        self.lbl_sep = ctk.CTkLabel(self.content, text="|", font=FONTS["main"], text_color="#40444b")
        self.lbl_timer = ctk.CTkLabel(self.content, text="00:00", font=FONTS["timer"], text_color=THEME["timer_idle"])
        self.lbl_deaths = ctk.CTkLabel(self.content, text="", font=FONTS["main"], text_color=THEME["death"])
        self.frm_mechanics = ctk.CTkFrame(self.content, fg_color="transparent")
        self.frm_tier = ctk.CTkFrame(self.content, fg_color="transparent")
        ctk.CTkLabel(self.frm_tier, text="T", font=FONTS["main"], text_color=THEME["timer_idle"]).pack(side="left", padx=(0,2))
        self.entry_tier = ctk.CTkEntry(self.frm_tier, textvariable=self.tier_var, width=30, height=20, 
                                       fg_color=THEME["input_bg"], border_width=0, 
                                       font=FONTS["main"], justify="center")
        self.entry_tier.pack(side="left")

        self._make_draggable(self.title_bar)
        self._make_draggable(self.lbl_title)
        self._apply_layout_standard()

    def _toggle_compact_mode(self):
        self.is_compact = not self.is_compact
        if self.is_compact: self._apply_layout_compact()
        else: self._apply_layout_standard()

    def _apply_layout_standard(self):
        self._unpack_all()
        self.lbl_title.pack(side="left", padx=5)
        self.geometry("340x50")
        self.frm_tier.pack(side="right", padx=2)
        self.lbl_map.pack(side="left", padx=(0, 5))
        self.lbl_sep.pack(side="left", padx=(0, 5))
        self.lbl_timer.pack(side="left", padx=(0, 5))
        self.lbl_deaths.pack(side="left", padx=(0, 5))
        self.frm_mechanics.pack(side="left", padx=(0, 5))

    def _apply_layout_compact(self):
        self._unpack_all()
        self.lbl_title.pack_forget()
        self.geometry("100x50")
        self.lbl_timer.pack(side="top", pady=0)

    def _unpack_all(self):
        self.lbl_map.pack_forget()
        self.lbl_sep.pack_forget()
        self.lbl_timer.pack_forget()
        self.lbl_deaths.pack_forget()
        self.frm_mechanics.pack_forget()
        self.frm_tier.pack_forget()

    def _make_draggable(self, widget):
        def start_move(e):
            self.x = e.x
            self.y = e.y
        def do_move(e):
            deltax = e.x - self.x
            deltay = e.y - self.y
            x = self.winfo_x() + deltax
            y = self.winfo_y() + deltay
            self.geometry(f"+{x}+{y}")
        widget.bind("<Button-1>", start_move)
        widget.bind("<B1-Motion>", do_move)

    def _monitor_focus_loop(self):
        if not self.running: return
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            if not hwnd: 
                self.after(200, self._monitor_focus_loop)
                return
            length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
            buf = ctypes.create_unicode_buffer(length + 1)
            ctypes.windll.user32.GetWindowTextW(hwnd, buf, length + 1)
            active_title = buf.value
            is_poe = "Path of Exile" in active_title
            is_tracker = f"PoE Map Tracker v{APP_VERSION}" in active_title
            
            if is_poe or is_tracker:
                if self.is_paused:
                    resume_duration = time.time() - self.pause_timestamp
                    if self.status == "running": self.start_time += resume_duration
                    elif self.status == "cooldown": self.cooldown_start += resume_duration
                    self.is_paused = False
                if self.state() == "withdrawn": self.deiconify()
                self.attributes("-topmost", True)
            else:
                if not self.is_paused and (self.status == "running" or self.status == "cooldown"):
                    self.is_paused = True
                    self.pause_timestamp = time.time()
                if self.state() != "withdrawn": self.withdraw()
        except Exception: pass
        self.after(200, self._monitor_focus_loop)

    def _monitor_game_process(self):
        if not self.running: return
        is_game_running = False
        try:
            output = subprocess.check_output('tasklist /FI "IMAGENAME eq PathOfExile*"', shell=True, creationflags=subprocess.CREATE_NO_WINDOW).decode('utf-8', errors='ignore')
            for proc in GAME_PROCESSES:
                if proc.lower() in output.lower():
                    is_game_running = True
                    break
        except: is_game_running = True

        if not is_game_running:
            self.game_not_found_count += 1
            countdown = 10 - (self.game_not_found_count * 2)
            if countdown <= 0:
                if hasattr(self, 'tray_icon'): self.tray_icon.stop()
                self.quit()
                return
            if not self.is_compact: self.lbl_map.configure(text=f"Game Closed! Closing {countdown}s...", text_color=THEME["death"])
        else: self.game_not_found_count = 0 
        self.after(2000, self._monitor_game_process)

    def _init_log_search(self):
        self.log_path = self._get_log_path()
        if self.log_path and os.path.exists(self.log_path):
            self.current_map = "No Hideout"
            threading.Thread(target=self._logic_loop, daemon=True).start()
            threading.Thread(target=self._monitor_log_file, daemon=True).start()
            self._update_gui_loop()
        else: self.lbl_map.configure(text="Log not found!", text_color=THEME["death"])

    def _get_log_path(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r") as f: return f.read().strip()
            except: pass
        candidates = [
            r"C:\Games\Steam\steamapps\common\Path of Exile\logs\Client.txt",
            r"C:\Program Files (x86)\Grinding Gear Games\Path of Exile\logs\Client.txt",
            r"C:\Program Files (x86)\Steam\steamapps\common\Path of Exile\logs\Client.txt",
            r"D:\SteamLibrary\steamapps\common\Path of Exile\logs\Client.txt",
            os.path.expanduser(r"~\Documents\My Games\Path of Exile\logs\Client.txt")
        ]
        for path in candidates:
            if os.path.exists(path):
                self._save_config(path)
                return path
        root = tk.Tk(); root.withdraw()
        messagebox.showinfo("Setup", "Please select 'Client.txt' in your Path of Exile logs folder.")
        selected = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")]); root.destroy()
        if selected: self._save_config(selected)
        return selected

    def _save_config(self, path):
        try: 
            with open(self.config_file, "w") as f: f.write(path)
        except: pass

    def _monitor_log_file(self):
        try:
            f = open(self.log_path, 'r', encoding='utf-8')
            f.seek(0, 2)
            while self.running:
                line = f.readline()
                if not line:
                    time.sleep(0.1)
                    continue
                self._process_log_line(line)
        except Exception: self.current_map = "Log Error"

    def _process_log_line(self, line):
        if " : You have entered " in line:
            try:
                zone = line.split(" : You have entered ")[1].strip().replace(".", "")
                self._handle_zone_change(zone)
            except: pass
        if " : You have been slain." in line and self.status == "running": self.deaths += 1
        if self.status == "running":
            if any(char in line for char in ["@", "#", "$", "%", "&"]): return
            for trigger, data in MECHANICS_CONFIG.items():
                if trigger in line: 
                    self._add_mechanic(data)
                    # Special Check for Zana/T16.5
                    if data["name"] == "Tier 16.5":
                        if "(T16.5)" not in self.current_map:
                            self.current_map += " (T16.5)"

    def _handle_zone_change(self, zone):
        is_safe = any(s in zone for s in SAFE_ZONES)
        
        # T17 CHECK (Append T17 tag automatically)
        if zone in TIER_17_MAPS and "(T17)" not in zone:
            zone += " (T17)"

        if is_safe:
            self.latest_safe_zone = zone # Store safe zone name
            if self.status == "running": 
                self.status = "cooldown"
                self.cooldown_start = time.time()
            elif self.status == "idle":
                self.current_map = zone # Update display instantly if idle
        else:
            if self.status == "running": 
                if self.current_map != zone:
                    self._prepare_run_data() 
                    self._start_run(zone)
            elif self.status == "cooldown":
                self.status = "running"
                self.current_map = zone
            else: self._start_run(zone)

    def _start_run(self, zone):
        self.status = "running"
        self.current_map = zone
        self.start_time = time.time()
        self.elapsed = 0
        self.deaths = 0 
        self.mechanics_found = []
        self.after(0, self._clear_mechanics_ui)

    def _clear_mechanics_ui(self):
        for widget in self.frm_mechanics.winfo_children(): widget.destroy()

    def _add_mechanic(self, mech_data):
        name = mech_data["name"]
        if name not in self.mechanics_found:
            self.mechanics_found.append(name)
            self.after(0, lambda: self._draw_mechanic_badge(mech_data))

    def _draw_mechanic_badge(self, data):
        badge = ctk.CTkLabel(self.frm_mechanics, text=data["badge"], fg_color=data["color"], text_color="black", font=FONTS["badge"], width=20, height=18, corner_radius=5)
        badge.pack(side="left", padx=1)

    def _prepare_run_data(self):
        if self.elapsed < 10: return
        run_data = {
            "date": datetime.now().strftime("%d/%m/%Y"),
            "time": datetime.now().strftime("%H:%M:%S"),
            "map": self.current_map,
            "tier": int(self.tier_var.get()) if self.tier_var.get().isdigit() else self.tier_var.get(),
            "duration": self._format_time(self.elapsed),
            "deaths": self.deaths,
            "mechanics": ", ".join(self.mechanics_found)
        }
        self.pending_runs.append(run_data)
        self.after(0, self._process_save_queue)

    def _process_save_queue(self):
        if not self.pending_runs: return
        try:
            self._ensure_history_folder()
            if not os.path.exists(self.excel_file): self._create_excel_file()
            wb = load_workbook(self.excel_file)
            for data in self.pending_runs: self._append_history(wb, data)
            self._update_statistics(wb)
            wb.save(self.excel_file)
            self.pending_runs.clear()
            self.lbl_map.configure(text="Saved!", text_color=THEME["text"])
        except Exception:
            self.lbl_map.configure(text="Excel Open! Retrying...", text_color=THEME["death"])
            self.after(5000, self._process_save_queue)

    def _create_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(["Date", "Time", "Map", "Tier", "Duration", "Deaths", "Mechanics"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F3136", end_color="2F3136", fill_type="solid")
        wb.create_sheet("Statistics")
        ws_stats = wb["Statistics"]
        ws_stats.append(["Map Name", "Count", "Last Tier", "Total Deaths", "Total Time", "", "Mechanic", "Count", "Total Time"])
        for cell in ws_stats[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            if cell.col_idx in [7, 8, 9]: cell.fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
        wb.save(self.excel_file)

    def _append_history(self, wb, data):
        ws = wb["History"]
        ws.append([data["date"], data["time"], data["map"], data["tier"], data["duration"], data["deaths"], data["mechanics"]])

    def _update_statistics(self, wb):
        ws_hist = wb["History"]
        ws_stats = wb["Statistics"]
        ws_stats.delete_rows(2, ws_stats.max_row + 1)
        maps_data = {}
        mechanics_data = {}
        for row in ws_hist.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5: continue
            m_name, m_tier = row[2], row[3]
            m_dur_str = row[4]
            m_deaths = row[5] if isinstance(row[5], int) else 0
            
            secs = 0
            try:
                parts = str(m_dur_str).split(':')
                if len(parts) == 2: secs = int(parts[0]) * 60 + int(parts[1])
                elif len(parts) == 3: secs = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
            except: pass

            if m_name:
                if m_name not in maps_data: maps_data[m_name] = {"count": 0, "tier": m_tier, "deaths": 0, "time": 0}
                maps_data[m_name]["count"] += 1
                maps_data[m_name]["deaths"] += m_deaths
                maps_data[m_name]["tier"] = m_tier
                maps_data[m_name]["time"] += secs
            m_mechs = row[6] if len(row) > 6 else ""
            if m_mechs:
                for mech in str(m_mechs).split(", "):
                    mech = mech.strip()
                    if mech: 
                        if mech not in mechanics_data: mechanics_data[mech] = {"count": 0, "time": 0}
                        mechanics_data[mech]["count"] += 1
                        mechanics_data[mech]["time"] += secs
        
        def fmt_dur(s):
            d = s // 86400; s %= 86400
            h = s // 3600; s %= 3600
            m = s // 60; s %= 60
            p = []
            if d > 0: p.append(f"{d}d")
            if h > 0: p.append(f"{h}h")
            if m > 0: p.append(f"{m}m")
            p.append(f"{s}s")
            return " ".join(p) if p else "0s"

        sorted_maps = sorted(maps_data.items(), key=lambda x: x[1]['count'], reverse=True)
        for idx, (name, data) in enumerate(sorted_maps):
            ws_stats.cell(row=idx+2, column=1, value=name)
            ws_stats.cell(row=idx+2, column=2, value=data['count'])
            ws_stats.cell(row=idx+2, column=3, value=data['tier'])
            ws_stats.cell(row=idx+2, column=4, value=data['deaths'])
            ws_stats.cell(row=idx+2, column=5, value=fmt_dur(data['time']))
        sorted_mechs = sorted(mechanics_data.items(), key=lambda x: x[1]['count'], reverse=True)
        for idx, (name, data) in enumerate(sorted_mechs):
            ws_stats.cell(row=idx+2, column=7, value=name)
            ws_stats.cell(row=idx+2, column=8, value=data['count'])
            ws_stats.cell(row=idx+2, column=9, value=fmt_dur(data['time']))
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['E'].width = 15
        ws_stats.column_dimensions['G'].width = 15
        ws_stats.column_dimensions['I'].width = 15

    def _logic_loop(self):
        while self.running:
            time.sleep(0.1)
            if self.is_paused: continue
            if self.status == "running": self.elapsed = time.time() - self.start_time
            elif self.status == "cooldown":
                remain = SAVE_DELAY - (time.time() - self.cooldown_start)
                if remain <= 0:
                    self._prepare_run_data()
                    self.status = "idle"
                    self.elapsed = 0
                    self.current_map = self.latest_safe_zone # Update display to current safe zone
                    self.deaths = 0
                    self.mechanics_found = []
                    
    def _update_gui_loop(self):
        if self.game_not_found_count == 0:
            if self.status == "cooldown":
                remain = int(SAVE_DELAY - (time.time() - self.cooldown_start))
                if self.is_compact: self.lbl_map.configure(text="") 
                else: self.lbl_map.configure(text=f"Saving: {remain}s", text_color="orange")
            else:
                if self.pending_runs:
                    if not self.is_compact: self.lbl_map.configure(text="Excel Open! Retrying...", text_color=THEME["death"])
                else:
                    if not self.is_compact:
                        name = self.current_map
                        if len(name) > 22: name = name[:20] + ".."
                        self.lbl_map.configure(text=name, text_color=THEME["text"])
        self.lbl_timer.configure(text=self._format_time(self.elapsed), text_color=THEME["accent"] if self.status == "running" else THEME["timer_idle"])
        self.lbl_deaths.configure(text=f"☠ {self.deaths}" if self.deaths > 0 else "")
        if self.status == "idle" and self.frm_mechanics.winfo_children():
            for widget in self.frm_mechanics.winfo_children(): widget.destroy()
        self.after(100, self._update_gui_loop)

    def _format_time(self, s): return f"{int(s//60):02d}:{int(s%60):02d}"

if __name__ == "__main__":
    app = PoEOverlay()
    app.mainloop()